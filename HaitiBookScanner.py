# isbntools documentation found at https://isbntools.readthedocs.io/en/latest/info.html
# Using this too: https://stackoverflow.com/questions/26360699/how-can-i-get-the-author-and-title-by-knowing-the-isbn-using-google-book-api
import sys
import openpyxl
import datetime
from isbntools.app import *

# Service set for Google Books
service = "wcat"

# Main menu
def main():
    print("Main Menu:")
    print("To register new books, type 'register'")
    print("To check in books, type 'check in'")
    print("To check out books, type 'check out'")
    print("To exit, type 'exit'")
    choice = input("")
    if choice in ["register", "Register", "REGISTER", "'register'"]:
        print("")
        register_book()
    elif choice in ["check in", "checkin", "Check In", "Check in", "CHECKIN", "CHECK IN", "CheckIn", "'check in'"]:
        print("")
        check_in()
    elif choice in ["check out", "checkout", "Check Out", "Check out", "CHECKOUT", "CHECK OUT", "CheckOut", "'check out'"]:
        print("")
        check_out()
    elif choice in ["exit", "Exit", "EXIT"]:
        exit()
    else:
        print("The inputted value is not an option. Try again.")
        print("")
        main()
# Allows for registration of books into database.
def register_book():
    inventory_workbook = openpyxl.load_workbook("BookDatabase.xlsx")
    book_inventory_sheet = inventory_workbook["Book Inventory"]
    book = input("Scan barcode to register book or type 'menu': ")
    if book == "menu":
        print("")
        main()
    else:
        isbn_list = []
        for row in book_inventory_sheet["C"]:
            isbn_list.append(row.value)
        if book in isbn_list:
            cell_row = (isbn_list.index(book) + 1)
            total_quantity = book_inventory_sheet["D" + str(cell_row)]
            in_stock = book_inventory_sheet["E" + str(cell_row)]
            total_quantity.value = (total_quantity.value + 1)
            in_stock.value = (in_stock.value + 1)
            print("At least one of this book already registered. Total quantity is now: " + str(total_quantity.value))
            print("")
            inventory_workbook.save("BookDatabase.xlsx")
            register_book()
        else:
            meta_dict = meta(book, service)
            authors_list = meta_dict["Authors"]
            authors = ",".join(authors_list)
            title = meta_dict["Title"]
            # Appends the info to the last column, and sets "Total Quantity" and "In Stock" to 1
            book_inventory_sheet.append([title, authors, book, 1, 1])
            print (title + " by " + authors + " added to database.")
            print("")
            inventory_workbook.save("BookDatabase.xlsx")
            register_book()

        
# Allows the library to scan books in once returned.
def check_in():
    time = datetime.datetime.now()
    current_date = time.strftime('%d-%m-%Y %H:%M:%S')
    inventory_workbook = openpyxl.load_workbook("BookDatabase.xlsx")
    book_history_sheet = inventory_workbook["Check Out-In"]
    book_inventory_sheet = inventory_workbook["Book Inventory"]
    book = input("Scan barcode to check in or type 'menu': ")
    print("")
    if book == "menu":
        main()
    else:
        inventory_isbn_list = []  # List of all ISBN numbers in Book Inventory sheet
        checked_out_list = []  # List of all ISBN numbers in Check In-Out sheet
        revised_checked_out_list = [] # List of all books that match the scanned ISBN that are checked out
        for row in book_inventory_sheet["C"]:
            inventory_isbn_list.append(row.value)
        if book in inventory_isbn_list:
            i = 0
            while i <= (len(inventory_isbn_list) - 1):
                name = book_history_sheet["D" + str(i + 2)].value
                isbn = book_history_sheet["C" + str(i + 2)].value
                row_location = i + 2
                checked_out_list.append([name, isbn, row_location])
                if isbn == book:
                    revised_checked_out_list.append([name, row_location])
                i += 1
            print ("Select the number next to the name of who is checking the book in:")
            n = 0
            while n <= (len(revised_checked_out_list) - 1):
                print(str(n + 1) + ": " + revised_checked_out_list[n][0])
                n += 1
            selected_number = int(input("Enter number next to name here: "))
            if selected_number <= n and selected_number > 0:
                print(str(selected_number))
                selected_person = revised_checked_out_list[selected_number - 1][1]
                book_history_sheet.delete_rows(selected_person, 1)
                cell_row = (inventory_isbn_list.index(book) + 1)
                in_stock = book_inventory_sheet["E" + str(cell_row)]
                in_stock.value = (in_stock.value + 1)
                inventory_workbook.save("BookDatabase.xlsx")
            else:
                print("")
                print("Selected number is not an option. Please try again.")
                print("")
                check_in()
            check_in()
        elif book not in inventory_isbn_list:
            print("ERROR: This book was never registered. Its ISBN number is not in the database.")
            print("")
            main()
    inventory_workbook.save("BookDatabase.xlsx")

# Allows the library to scan books when checked out.
def check_out():
    time = datetime.datetime.now()
    current_date = time.strftime('%d-%m-%Y %H:%M:%S')
    inventory_workbook = openpyxl.load_workbook("BookDatabase.xlsx")
    book_history_sheet = inventory_workbook["Check Out-In"]
    book_inventory_sheet = inventory_workbook["Book Inventory"]
    book = input("Scan barcode to check out or type 'menu': ")
    if book == "menu":
        print("")
        main()
    else:
        isbn_list = []
        for row in book_inventory_sheet["C"]:
            isbn_list.append(row.value)
        if book in isbn_list:
            cell_row = (isbn_list.index(book) + 1)
            in_stock = book_inventory_sheet["E" + str(cell_row)]
            meta_dict = meta(book, service)
            authors_list = meta_dict["Authors"]
            authors = ",".join(authors_list)
            title = meta_dict["Title"]
            if in_stock.value <= 0:
                print("ERROR: The database claims that there are 0 books left in stock. Did you mean to check in?")
                print("")
                check_out()
            else:
                checked_out_by = input("Enter the name of who is checking out the book: ")
                in_stock.value = (in_stock.value - 1)
                book_history_sheet.append([title, authors, book, checked_out_by, current_date])
                print(title + " successfully checked out to " + checked_out_by + ". Remaining copies of this book: " + str(in_stock.value))
                print("")
                inventory_workbook.save("BookDatabase.xlsx")
                check_out()
        else:
            print("ERROR: This book was never registered. Its ISBN number is not in the database.")
            check_out()


main()
